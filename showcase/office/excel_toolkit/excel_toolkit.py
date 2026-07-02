"""
[Showcase / Office / 通用]
Skill: Excel 读写工具

适用场景：
  批量导出报表数据到 Excel，或读取别人发来的 Excel 提取数据做后续处理。
  纯文件格式操作（openpyxl 直接读写 .xlsx 结构），不需要打开 Excel 应用，
  不需要屏幕，服务器上也能跑。

用法：
  # 读取 Excel，输出为 JSON
  python run.py showcase/office/excel_toolkit/excel_toolkit -- --read data.xlsx

  # 读取指定 sheet
  python run.py showcase/office/excel_toolkit/excel_toolkit -- --read data.xlsx --sheet "Sheet2"

  # 写入数据到新 Excel（--data 是 JSON 数组，每项是一行的字段字典）
  python run.py showcase/office/excel_toolkit/excel_toolkit -- \\
    --write output.xlsx --data '[{"姓名":"张三","分数":90},{"姓名":"李四","分数":85}]'
"""

import argparse
import json
import sys

from openpyxl import Workbook, load_workbook

from core.logger import SkillLogger


def read_excel(path: str, sheet: str = "") -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


def write_excel(path: str, data: list[dict], sheet: str = "Sheet1") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    if data:
        headers = list(data[0].keys())
        ws.append(headers)
        for row in data:
            ws.append([row.get(h, "") for h in headers])
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Excel 读写工具")
    parser.add_argument("--read", default="", help="读取的 Excel 文件路径")
    parser.add_argument("--write", default="", help="写入的 Excel 文件路径")
    parser.add_argument("--sheet", default="", help="指定 sheet 名称（读取时不填默认取第一个，写入时不填默认 Sheet1）")
    parser.add_argument("--data", default="", help="写入用的 JSON 数组数据（配合 --write）")
    try:
        sep = sys.argv.index("--")
        argv = sys.argv[sep + 1:]
    except ValueError:
        argv = sys.argv[1:]
    args = parser.parse_args(argv)

    log = SkillLogger("office/excel_toolkit")

    if args.read:
        log.step(f"读取 {args.read}")
        rows = read_excel(args.read, args.sheet)
        print(json.dumps(rows, ensure_ascii=False, indent=2))
        log.finish({"rows": len(rows)})
        return rows

    if args.write:
        data = json.loads(args.data) if args.data else []
        log.step(f"写入 {len(data)} 行到 {args.write}")
        write_excel(args.write, data, args.sheet or "Sheet1")
        print(f"已保存到 {args.write}")
        log.finish({"rows": len(data), "output": args.write})
        return {"output": args.write, "rows": len(data)}

    parser.print_help()


if __name__ == "__main__":
    main()
